"""Add an 'Occupation' column to demo_data.xlsx's scoreboard.

Each player gets a value chosen randomly from {Tech Arch, Developer, DM}. The
choice is seeded so the committed workbook is reproducible (re-running yields
the same assignment). The scoreboard named range grows from D3:F9 to D3:G9 so
edjas reads the new column. Run with:
    uv run python scripts/add_occupation.py
"""
import os
import random

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK = os.path.join(ROOT, "data", "demo_data.xlsx")
OCCUPATIONS = ["Tech Arch", "Developer", "DM"]

random.seed(20260625)

wb = openpyxl.load_workbook(WORKBOOK)
ws = wb["Sheet1"]

ws["G3"] = "Occupation"
for row in range(4, 10):  # the six players live in rows 4..9
    ws[f"G{row}"] = random.choice(OCCUPATIONS)

wb.defined_names["scoreboard"].value = "Sheet1!$D$3:$G$9"

wb.save(WORKBOOK)
assigned = [ws[f"G{r}"].value for r in range(4, 10)]
print(f"added Occupation column {assigned}; scoreboard now {wb.defined_names['scoreboard'].value}")
