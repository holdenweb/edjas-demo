"""Nest the demo's version into a {number} dict.

Sets Sheet1!B4 to '{version}' and adds a one-row 'version' named range
(number | 0.1.2). This relies on the edjas flatten fix, which lets a
single-entry dict read correctly instead of being flattened to a vector.
Idempotent; demo_data.xlsx is modified in place (still on Sheet1). Run with:
    uv run python scripts/nest_version.py
"""
import os

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK = os.path.join(ROOT, "data", "demo_data.xlsx")

wb = openpyxl.load_workbook(WORKBOOK)
ws = wb["Sheet1"]

# Point the Parameters 'version' value at a sub-range, and write that range.
ws["B4"] = "{version}"
ws["A11"], ws["B11"] = "number", "0.1.2"

if "version" in wb.defined_names:
    wb.defined_names["version"].value = "Sheet1!$A$11:$B$11"
else:
    wb.defined_names.add(DefinedName("version", attr_text="Sheet1!$A$11:$B$11"))

wb.save(WORKBOOK)
print(f"nested version: B4={ws['B4'].value!r}, version -> {wb.defined_names['version'].value}")
