"""Build levels_sample.xlsx for the 'levels of detail' demo.

Starts from a copy of demo_data.xlsx (so it stays compatible with hubris) and
adds a single 'detail' key/value to the Parameters range, then extends that
named range to include the new row. The result is a workbook whose author has
chosen the report's opening altitude — no template or code change needed.

The original demo_data.xlsx is left untouched. Run with:
    uv run python scripts/make_levels_sample.py
"""
import os

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "demo_data.xlsx")
DST = os.path.join(ROOT, "levels_sample.xlsx")

wb = openpyxl.load_workbook(SRC)
ws = wb["Sheet1"]

# A6/B6 are free (the scoreboard lives in columns D:F), so the new key/value
# pair slots straight into the Parameters block.
ws["A6"] = "detail"
ws["B6"] = "executive"

# Grow the Parameters named range from A3:B5 to A3:B6 so hubris reads the cell.
wb.defined_names["Parameters"].value = "Sheet1!$A$3:$B$6"

wb.save(DST)
print(f"wrote {DST} — Parameters now {wb.defined_names['Parameters'].value}, detail=executive")
